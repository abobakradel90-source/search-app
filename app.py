import streamlit as st
import streamlit.components.v1 as components
import chromadb
from transformers import CLIPProcessor, CLIPModel
from PIL import Image, ImageOps, ImageChops
import torch
import os
import zipfile
import urllib.request
import pandas as pd
import shutil
import base64
from streamlit_cropper import st_cropper
import io
import datetime
import json
import re
import time
import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ==============================================================================
# 🌐 الرابط السحابي الدائم لقاعدة البيانات (Firebase Realtime Database)
# ==============================================================================
FIREBASE_DB_URL = "https://edstore-2be25-default-rtdb.firebaseio.com/"

# --- 1. إعدادات الصفحة وبوابة الدخول ---
st.set_page_config(
    page_title="ED STORE | بوابة النظام",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="collapsed"
)

USERS = {
    "abobakr": "admin2026",    
    "mohamed": "123456",       
    "ahmed": "edstore"         
}

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False
if 'current_user' not in st.session_state:
    st.session_state.current_user = ""

# ==========================================
# 🌐 إدارة البيانات السحابية والمحلية المزدوجة
# ==========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SHARED_INV_FILE = os.path.join(BASE_DIR, "shared_inventory.json")
MASTER_DB_FILE = os.path.join(BASE_DIR, "master_database.csv")
HISTORY_INV_FILE = os.path.join(BASE_DIR, "inventory_history.json")
SHARED_SALES_FILE = os.path.join(BASE_DIR, "shared_sales.json")
HISTORY_SALES_FILE = os.path.join(BASE_DIR, "sales_history.json")

def cloud_get(endpoint, default_val):
    if FIREBASE_DB_URL and not "edstore-default" in FIREBASE_DB_URL:
        try:
            url = f"{FIREBASE_DB_URL.rstrip('/')}/{endpoint}.json"
            r = requests.get(url, timeout=3.5)
            if r.status_code == 200 and r.json() is not None:
                return r.json()
        except Exception:
            pass
    return default_val

def cloud_put(endpoint, data):
    if FIREBASE_DB_URL and not "edstore-default" in FIREBASE_DB_URL:
        try:
            url = f"{FIREBASE_DB_URL.rstrip('/')}/{endpoint}.json"
            requests.put(url, json=data, timeout=3.5)
        except Exception:
            pass

def load_json(file_path, default_data):
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return default_data

def save_json(file_path, data):
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except Exception:
        pass

def load_shared_inventory():
    default_inv = {"is_active": False, "name": "", "reason": "", "date": "", "scanned_items": {}}
    cloud_data = cloud_get("active_inventory", None)
    if cloud_data is not None:
        save_json(SHARED_INV_FILE, cloud_data)
        return cloud_data
    return load_json(SHARED_INV_FILE, default_inv)

def save_shared_inventory(data):
    save_json(SHARED_INV_FILE, data)
    cloud_put("active_inventory", data)

def load_inv_history():
    cloud_data = cloud_get("inventory_history", None)
    if cloud_data is not None:
        save_json(HISTORY_INV_FILE, cloud_data)
        return cloud_data
    return load_json(HISTORY_INV_FILE, [])

def save_to_inv_history(record):
    history = load_inv_history()
    history.append(record)
    save_json(HISTORY_INV_FILE, history)
    cloud_put("inventory_history", history)

def load_shared_sales():
    default_sales = {"is_active": False, "name": "", "date": "", "invoices": [], "deductions": {}}
    cloud_data = cloud_get("active_sales", None)
    if cloud_data is not None:
        save_json(SHARED_SALES_FILE, cloud_data)
        return cloud_data
    return load_json(SHARED_SALES_FILE, default_sales)

def save_shared_sales(data):
    save_json(SHARED_SALES_FILE, data)
    cloud_put("active_sales", data)

def load_sales_history():
    cloud_data = cloud_get("sales_history", None)
    if cloud_data is not None:
        save_json(HISTORY_SALES_FILE, cloud_data)
        return cloud_data
    return load_json(HISTORY_SALES_FILE, [])

def save_to_sales_history(record):
    history = load_sales_history()
    history.append(record)
    save_json(HISTORY_SALES_FILE, history)
    cloud_put("sales_history", history)

def get_image_base64(img_path):
    try:
        with open(img_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode('utf-8')
    except Exception:
        return ""

def process_shoe_image(img_path, target_w=280, target_h=180, quality=88):
    try:
        with Image.open(img_path) as img:
            img_rgb = img.convert("RGB")
            w, h = img_rgb.size
            crop_box = None
            
            if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                try:
                    alpha = img.convert('RGBA').split()[-1]
                    a_box = alpha.point(lambda p: 255 if p > 15 else 0).getbbox()
                    if a_box and (a_box[2] - a_box[0] > 20) and (a_box[3] - a_box[1] > 20):
                        if a_box != (0, 0, w, h):
                            crop_box = a_box
                except Exception:
                    pass
            
            if not crop_box:
                corners = [
                    img_rgb.getpixel((0, 0)),
                    img_rgb.getpixel((w - 1, 0)),
                    img_rgb.getpixel((0, h - 1)),
                    img_rgb.getpixel((w - 1, h - 1))
                ]
                avg_bg = (
                    int(sum(c[0] for c in corners) / 4),
                    int(sum(c[1] for c in corners) / 4),
                    int(sum(c[2] for c in corners) / 4)
                )
                
                for tol in [12, 22, 35, 50, 65]:
                    bg_img = Image.new('RGB', (w, h), avg_bg)
                    diff = ImageChops.difference(img_rgb, bg_img).convert('L')
                    mask = diff.point(lambda p: 255 if p > tol else 0)
                    bbox = mask.getbbox()
                    if bbox:
                        bw, bh = bbox[2] - bbox[0], bbox[3] - bbox[1]
                        if bw < w * 0.98 or bh < h * 0.98:
                            crop_box = bbox
                            break
                        crop_box = bbox
            
            if crop_box:
                pad_x = max(2, int((crop_box[2] - crop_box[0]) * 0.02))
                pad_y = max(2, int((crop_box[3] - crop_box[1]) * 0.02))
                safe_box = (
                    max(0, crop_box[0] - pad_x),
                    max(0, crop_box[1] - pad_y),
                    min(w, crop_box[2] + pad_x),
                    min(h, crop_box[3] + pad_y)
                )
                shoe_cropped = img_rgb.crop(safe_box)
            else:
                shoe_cropped = img_rgb
                
            sw, sh = shoe_cropped.size
            scale = min((target_w - 12) / sw, (target_h - 12) / sh)
            nw = max(1, int(sw * scale))
            nh = max(1, int(sh * scale))
            
            resized_shoe = shoe_cropped.resize((nw, nh), Image.Resampling.LANCZOS)
            canvas = Image.new("RGB", (target_w, target_h), (255, 255, 255))
            offset = ((target_w - nw) // 2, (target_h - nh) // 2)
            canvas.paste(resized_shoe, offset)
            
            buf = io.BytesIO()
            canvas.save(buf, format="JPEG", quality=quality, optimize=True)
            buf.seek(0)
            return buf
    except Exception:
        return None

def get_thumbnail_base64(img_path):
    buf = process_shoe_image(img_path, target_w=200, target_h=130, quality=85)
    if buf:
        return f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"
    return None

def generate_catalog_excel(catalog_items):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "الكاتالوج"
    try: ws.sheet_view.rightToLeft = True
    except: pass
    
    headers = ["صورة المنتج", "كود الصنف", "اسم الصنف", "سعر القطعة (ج.م)", "الرصيد الدفتري (قبل البيع)", "كمية المبيعات بالوردية", "الرصيد اللحظي المتاح"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1C65A6", end_color="1C65A6", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
    )
    
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    ws.row_dimensions[1].height = 30
    
    ws.column_dimensions['A'].width = 23
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    
    for row_idx, item in enumerate(catalog_items, start=2):
        ws.row_dimensions[row_idx].height = 105
        ws.cell(row=row_idx, column=1, value="")
        
        c_code = item.get("كود الصنف", "")
        c_name = item.get("اسم الصنف", "")
        ws.cell(row=row_idx, column=2, value=str(c_code if pd.notna(c_code) else "")).alignment = center_align
        ws.cell(row=row_idx, column=3, value=str(c_name if pd.notna(c_name) else "")).alignment = center_align
        
        try: p_val = float(item.get("سعر القطعة (ج.م)", 0.0))
        except: p_val = 0.0
        ws.cell(row=row_idx, column=4, value=p_val if pd.notna(p_val) else 0.0).alignment = center_align
        
        try: s_val = float(item.get("الرصيد الدفتري (قبل البيع)", 0.0))
        except: s_val = 0.0
        ws.cell(row=row_idx, column=5, value=s_val if pd.notna(s_val) else 0.0).alignment = center_align
        
        try: m_val = float(item.get("كمية المبيعات بالوردية", 0.0))
        except: m_val = 0.0
        ws.cell(row=row_idx, column=6, value=m_val if pd.notna(m_val) else 0.0).alignment = center_align
        
        try: a_val = float(item.get("الرصيد اللحظي المتاح", 0.0))
        except: a_val = 0.0
        ws.cell(row=row_idx, column=7, value=a_val if pd.notna(a_val) else 0.0).alignment = center_align
        
        for col_num in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_num).border = thin_border
            
        img_path = item.get("img_path")
        if img_path and isinstance(img_path, str) and os.path.exists(img_path):
            try:
                img_buf = process_shoe_image(img_path, target_w=280, target_h=180, quality=88)
                if img_buf:
                    xl_img = OpenpyxlImage(img_buf)
                    xl_img.width = 145
                    xl_img.height = 95
                    ws.add_image(xl_img, f"A{row_idx}")
            except Exception:
                pass
                
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

logo_base64 = get_image_base64("edstore.jpg")

# --- 2. الهوية البصرية ---
st.markdown(f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Tajawal:wght@400;500;700;800;900&display=swap');
        :root {{ --primary-color: #1C65A6 !important; }}
        html, body, [class*="css"] {{ font-family: 'Tajawal', sans-serif !important; direction: rtl; }}
        .stApp {{ background-color: #F0F4F8; }}
        #MainMenu {{ visibility: hidden; }}
        footer {{ visibility: hidden; }}
        header[data-testid="stHeader"] {{ background: transparent !important; }}
        [data-testid="stToolbar"] {{ display: none !important; }}
        [data-testid="stSidebarResizerRoot"], [data-testid="stSidebarCollapseButton"] {{ display: none !important; }}
        .block-container {{ padding-top: 1.5rem !important; padding-bottom: 5rem !important; max-width: 1300px; }}
        
        .stTabs [data-baseweb="tab-list"] {{
            gap: 12px;
            justify-content: center;
            background: transparent !important;
            padding: 5px;
            margin-bottom: 25px;
            border-bottom: none !important;
            flex-wrap: wrap;
        }}
        .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] {{ display: none !important; }}
        
        .stTabs [data-baseweb="tab"] {{
            background-color: #FFFFFF !important;
            border: 2px solid #CBD5E1 !important;
            border-radius: 12px !important;
            padding: 9px 20px !important;
            font-size: 16px !important;
            font-weight: 800 !important;
            color: #475569 !important;
            box-shadow: 0 2px 5px rgba(0, 0, 0, 0.04) !important;
            transition: all 0.2s ease !important;
        }}
        .stTabs [data-baseweb="tab"]:hover {{
            border-color: #1C65A6 !important;
            color: #1C65A6 !important;
            background-color: #F8FAFC !important;
            transform: translateY(-1px);
        }}
        .stTabs [data-baseweb="tab"][aria-selected="true"] {{
            background-color: #1C65A6 !important;
            border: 2px solid #1C65A6 !important;
            color: #FFFFFF !important;
            box-shadow: 0 4px 12px rgba(28, 101, 166, 0.25) !important;
        }}
        
        div[data-baseweb="input"], div[data-baseweb="base-input"], div[data-baseweb="select"] > div {{
            background-color: #FFFFFF !important;
            border: 2px solid #94A3B8 !important;
            border-radius: 10px !important;
            box-shadow: 0 1px 3px rgba(0,0,0,0.04) !important;
        }}
        div[data-baseweb="input"]:focus-within, div[data-baseweb="base-input"]:focus-within, div[data-baseweb="select"]:focus-within > div {{
            border-color: #1C65A6 !important;
            box-shadow: 0 0 0 3px rgba(28, 101, 166, 0.2) !important;
            outline: none !important;
        }}
        
        .login-card {{
            background: #FFFFFF;
            padding: 40px 35px;
            border-radius: 24px;
            box-shadow: 0 20px 45px rgba(28, 101, 166, 0.12);
            max-width: 440px;
            margin: 40px auto;
            text-align: center;
            border: 1px solid #E2E8F0;
            border-top: 6px solid #1C65A6;
        }}
        .login-logo-wrap {{ display: flex; justify-content: center; align-items: center; margin-bottom: 15px; }}
        .login-logo-img {{ height: 90px; width: 90px; object-fit: cover; border-radius: 50%; border: 3px solid #E8F0F8; background: #FFFFFF; padding: 3px; }}
        .login-title {{ color: #0F2942; font-weight: 900; font-size: 26px; margin-bottom: 6px; }}
        .login-subtitle {{ color: #64748B; font-size: 14px; margin-bottom: 25px; }}
        
        .brand-navbar {{ background: linear-gradient(90deg, #1C65A6 0%, #144A7A 100%); padding: 15px 20px; display: flex; align-items: center; justify-content: center; color: white; width: 100%; margin-top: -20px; margin-bottom: 15px; border-bottom-left-radius: 20px; border-bottom-right-radius: 20px; box-shadow: 0 4px 15px rgba(28, 101, 166, 0.3); }}
        .brand-navbar img {{ height: 55px; margin-left: 15px; border-radius: 8px; background-color: white; padding: 2px; }}
        .brand-navbar h1 {{ color: white; margin: 0; font-weight: 900; font-size: 2.2rem; }}
        .stButton > button {{ background-color: #1C65A6 !important; color: white !important; border-radius: 10px; border: none; padding: 9px 18px; font-weight: 800; font-size: 15px; width: 100%; box-shadow: 0 4px 10px rgba(28, 101, 166, 0.2); }}
        .stButton > button:hover {{ background-color: #144A7A !important; }}
        .logout-btn > button {{ background-color: #EF4444 !important; font-size: 14px !important; padding: 5px 15px !important; width: auto !important; }}
        
        .product-card {{ background: white; border-radius: 15px; padding: 20px; box-shadow: 0 4px 10px rgba(0,0,0,0.05); display: flex; align-items: center; gap: 25px; margin-bottom: 20px; border: 1px solid #E1E8F0; border-right: 6px solid #1C65A6; }}
        .product-img {{ width: 120px; height: 120px; object-fit: contain; border-radius: 10px; background-color: #F8FAFC; padding: 5px; border: 1px solid #E1E8F0; }}
        .code-badge {{ background-color: #E8F0F8; color: #1C65A6; padding: 4px 12px; border-radius: 20px; font-size: 13px; font-weight: 800; display: inline-block; margin-bottom: 8px; }}
        .product-title {{ font-size: 18px; color: #0F172A; font-weight: 800; margin: 0 0 6px 0; }}
        .stock-badge {{ display: inline-block; margin-top: 5px; padding: 4px 10px; border-radius: 8px; font-weight: 700; font-size: 14px; }}
        .in-stock {{ background-color: #ECFDF5; color: #059669; border: 1px solid #A7F3D0; }}
        .out-of-stock {{ background-color: #FEF2F2; color: #DC2626; border: 1px solid #FECACA; }}
        .metric-card {{ background: white; padding: 15px; border-radius: 12px; text-align: center; box-shadow: 0 4px 10px rgba(0,0,0,0.05); border-bottom: 4px solid #1C65A6; }}
        .metric-title {{ color: #64748B; font-size: 15px; font-weight: 700; margin-bottom: 5px; }}
        .metric-value {{ color: #0F172A; font-size: 28px; font-weight: 900; }}
        .inv-active-bar {{ background: linear-gradient(90deg, #10B981, #059669); color: white; padding: 12px 20px; border-radius: 10px; margin-bottom: 20px; font-weight: bold; }}
        .sales-active-bar {{ background: linear-gradient(90deg, #F59E0B, #D97706); color: white; padding: 12px 20px; border-radius: 10px; margin-bottom: 20px; font-weight: bold; }}
        .footer {{ position: fixed; bottom: 0; left: 0; width: 100%; background-color: #144A7A; color: white; text-align: center; padding: 10px; font-weight: 500; z-index: 999; }}
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 🛑 بوابة تسجيل الدخول
# ==========================================
if not st.session_state.logged_in:
    st.markdown('<div class="login-card">', unsafe_allow_html=True)
    if logo_base64:
        st.markdown(f'<div class="login-logo-wrap"><img src="data:image/jpeg;base64,{logo_base64}" class="login-logo-img"></div>', unsafe_allow_html=True)
    st.markdown('<div class="login-title">ED STORE</div>', unsafe_allow_html=True)
    st.markdown('<div class="login-subtitle">نظام إدارة المخزون والمبيعات الذكي</div>', unsafe_allow_html=True)
    
    username = st.text_input("👤 اسم المستخدم", placeholder="ادخل اسم المستخدم", key="login_u")
    password = st.text_input("🔑 كلمة المرور", placeholder="ادخل كلمة المرور", type="password", key="login_p")
    
    if st.button("تسجيل الدخول 🚀", key="login_btn"):
        clean_user = str(username).strip().lower()
        clean_pass = str(password).strip()
        if clean_user in USERS and USERS[clean_user] == clean_pass:
            st.session_state.logged_in = True
            st.session_state.current_user = clean_user
            st.rerun()
        else:
            st.error("❌ اسم المستخدم أو كلمة المرور غير صحيحة.")
    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ==========================================
# ✅ محتوى النظام الرئيسي
# ==========================================
if logo_base64: st.markdown(f'<div class="brand-navbar"><img src="data:image/jpeg;base64,{logo_base64}"><h1>ED STORE</h1></div>', unsafe_allow_html=True)
else: st.markdown('<div class="brand-navbar"><h1>ED STORE</h1></div>', unsafe_allow_html=True)

col_welc, col_out = st.columns([4, 1])
with col_welc: st.markdown(f"<h4 style='color:#1C65A6; margin:0;'>👤 مرحباً بك: <b>{st.session_state.current_user}</b></h4>", unsafe_allow_html=True)
with col_out:
    st.markdown('<div class="logout-btn">', unsafe_allow_html=True)
    if st.button("تسجيل خروج 🚪", key="logout_btn"):
        st.session_state.logged_in = False
        st.session_state.current_user = ""
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

if st.session_state.current_user == "abobakr":
    with st.expander("⚙️ لوحة الإدارة: رفع وتحديث شيت الأسعار والأرصدة (Excel / CSV)", expanded=False):
        c_up1, c_up2 = st.columns([3, 1])
        with c_up1:
            new_db = st.file_uploader("اختر شيت الإكسيل المحدّث:", type=['csv', 'xlsx'], key="admin_main_uploader")
        with c_up2:
            st.write("")
            st.write("")
            if new_db is not None:
                if st.button("💾 تطبيق وتحديث الداتا", key="apply_db_btn"):
                    try:
                        if new_db.name.endswith('.csv'): d = pd.read_csv(new_db, encoding='utf-8-sig', sep=None, engine='python')
                        else: d = pd.read_excel(new_db)
                        d.to_csv(MASTER_DB_FILE, index=False, encoding='utf-8-sig')
                        st.success("✅ تم تحديث قاعدة البيانات بنجاح!")
                        time.sleep(0.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"حدث خطأ أثناء القراءة: {e}")

st.markdown("<br>", unsafe_allow_html=True)

# --- محرك البحث الذكي ---
@st.cache_resource
def download_chroma_db():
    zip_p = os.path.join(BASE_DIR, "chroma_db.zip")
    ext_p = os.path.join(BASE_DIR, "chroma_db")
    mark_f = os.path.join(ext_p, "fashion_clip_v3.txt")
    download_url = "https://github.com/abobakradel90-source/search-app/releases/download/v1.0/chroma_db.zip"
    if os.path.exists(ext_p) and not os.path.exists(mark_f): shutil.rmtree(ext_p)
    if not os.path.exists(ext_p):
        with st.spinner('📦 جاري تهيئة مستودع البيانات الذكي...'):
            try:
                urllib.request.urlretrieve(download_url, zip_p)
                with zipfile.ZipFile(zip_p, 'r') as zip_ref: zip_ref.extractall(BASE_DIR)
                if os.path.exists(zip_p): os.remove(zip_p)
                with open(mark_f, 'w') as f: f.write("done")
            except Exception: pass

@st.cache_resource
def load_vision():
    download_chroma_db()
    model_id = "patrickjohncyh/fashion-clip"
    return CLIPModel.from_pretrained(model_id), CLIPProcessor.from_pretrained(model_id), chromadb.PersistentClient(path=os.path.join(BASE_DIR, "chroma_db")).get_collection(name="products_collection")

try: model, processor, collection = load_vision()
except Exception as e: st.error(f"⚠️ خطأ محرك الذكاء الاصطناعي: {e}")

def get_image_embedding(image):
    image = ImageOps.autocontrast(image.convert("RGB"), cutoff=1)
    inputs = processor(images=image, return_tensors="pt")
    with torch.no_grad():
        features = model.get_image_features(**inputs)
        if hasattr(features, 'image_embeds'): features = features.image_embeds
        elif hasattr(features, 'pooler_output'): features = features.pooler_output
        elif not isinstance(features, torch.Tensor): features = features[0]
        emb = features.squeeze().cpu().detach().numpy().flatten().tolist()
        return [float(x) for x in emb]

def get_color_histogram(image):
    img = image.convert("RGB").crop((image.size[0]*0.15, image.size[1]*0.15, image.size[0]*0.85, image.size[1]*0.85))
    hist = img.histogram(); total = sum(hist) / 3
    return [x / (total if total > 0 else 1) for x in hist]

def compare_histograms(h1, h2): return sum(abs(a - b) for a, b in zip(h1, h2))

# ==========================================
# 🚀 محرك استخراج البيانات الموحد
# ==========================================
system_inventory = {}

def parse_val(val):
    try:
        m = re.search(r'-?\d+(\.\d+)?', str(val).replace(',', '').strip())
        return float(m.group()) if m else 0.0
    except: return 0.0

def process_df(df):
    if df is None or df.empty: return
    cols_map = {c: str(c).lower().strip() for c in df.columns}
    code_col, name_col, stock_col, price_col = None, None, None, None
    
    for orig, low in cols_map.items():
        if any(k in low for k in ['كود الصنف', 'كود', 'code', 'باركود', 'barcode', 'item_code', 'رمز', 'رقم الصنف']):
            code_col = orig; break
            
    for orig, low in cols_map.items():
        if orig != code_col and any(k in low for k in ['اسم الصنف', 'اسم', 'name', 'صنف', 'item', 'title', 'موديل', 'model', 'الوصف', 'description']):
            name_col = orig; break
            
    for orig, low in cols_map.items():
        if orig not in [code_col, name_col] and any(k in low for k in ['الرصيد', 'رصيد', 'stock', 'الكمية', 'كمية', 'الكميه', 'كميه', 'qty', 'عدد', 'المخزون', 'المتاح']):
            stock_col = orig; break
            
    for orig, low in cols_map.items():
        if orig not in [code_col, name_col, stock_col] and any(k in low for k in ['سعر الجملة', 'سعر القطعة', 'سعر', 'price', 'ثمن', 'جملة', 'بيع', 'قيمة', 'قطاعي']):
            price_col = orig; break
            
    if not code_col and len(df.columns) > 0: code_col = df.columns[0]
    if not name_col and len(df.columns) > 1: name_col = df.columns[1]
    if not stock_col and len(df.columns) > 2: stock_col = df.columns[2]
    if not price_col and len(df.columns) > 3: price_col = df.columns[3]
    if not price_col and len(df.columns) > 4: price_col = df.columns[4]

    for _, row in df.iterrows():
        raw_code = row.get(code_col, "") if code_col else ""
        if pd.isna(raw_code) or str(raw_code).strip().lower() in ['nan', 'none', '']: continue
        
        p_code = str(raw_code).strip()
        if p_code.endswith('.0'): p_code = p_code[:-2]
        p_code = p_code.upper()
        if not p_code: continue
        
        p_name = str(row.get(name_col, "بدون اسم")).strip() if name_col else "بدون اسم"
        if pd.isna(p_name) or p_name.lower() in ['nan', 'none', '']: p_name = "بدون اسم"
        
        p_stock = parse_val(row.get(stock_col, 0)) if stock_col else 0.0
        p_price = parse_val(row.get(price_col, 0)) if price_col else 0.0
        
        if p_price == 0.0:
            for c in df.columns:
                if c != code_col and any(k in str(c).lower().strip() for k in ['price', 'سعر', 'ثمن', 'جملة']):
                    t_p = parse_val(row[c])
                    if t_p > 0: p_price = t_p; break
                    
        system_inventory[p_code] = {'name': p_name, 'sys_stock': p_stock, 'price': p_price}

if os.path.exists(MASTER_DB_FILE):
    try:
        df_m = pd.read_csv(MASTER_DB_FILE, encoding='utf-8-sig', sep=None, engine='python')
        process_df(df_m)
    except: pass
else:
    try:
        df_p = pd.read_csv(os.path.join(BASE_DIR, 'products.csv'), encoding='utf-8-sig', sep=None, engine='python')
        process_df(df_p)
    except: pass

def match_product_code(raw_code):
    if not raw_code:
        return None
    code_clean = str(raw_code).strip().upper()
    if code_clean in system_inventory:
        return code_clean
    
    stripped = code_clean.lstrip('0')
    for k in system_inventory.keys():
        if k == code_clean or (stripped and k.lstrip('0') == stripped):
            return k
    return None

def render_product_card(p_code, p_name, p_stock, p_price=None, is_sales=False):
    img_html = '<div class="product-img" style="display:flex; align-items:center; justify-content:center; color:#999; font-size:12px;">بدون صورة</div>'
    for ext in ['.jpg', '.jpeg', '.png', '.JPG']:
        img_path = os.path.join(BASE_DIR, "compressed_images", f"{p_code}{ext}")
        if os.path.exists(img_path):
            img_html = f'<img src="data:image/jpeg;base64,{get_image_base64(img_path)}" class="product-img">'
            break
            
    is_out = float(p_stock) <= 0
    stock_class = "stock-badge out-of-stock" if is_out else "stock-badge in-stock"
    price_str = f" &nbsp;|&nbsp; 💰 السعر: {p_price} ج.م" if p_price and float(p_price) > 0 else " &nbsp;|&nbsp; ⚠️ السعر غير مسجل"
    stock_text = f"🛒 الرصيد المتاح: {p_stock}{price_str}" if is_sales else f"📦 الرصيد الدفتري: {p_stock}"

    st.markdown(f"""<div class="product-card">
        {img_html}
        <div style="flex-grow: 1;">
            <div class="code-badge">الكود: {p_code}</div>
            <h3 class="product-title">{p_name}</h3>
            <div class="{stock_class}">{stock_text}</div>
        </div>
    </div>""", unsafe_allow_html=True)

# 🌟 التبويبات الرئيسية 🌟
tabs = st.tabs(["🔍 محرك البحث الذكي", "📦 الجرد التشاركي", "🛒 فواتير الجملة", "📖 الكاتالوج", "📈 لوحة تحكم الإدارة"])
main_tab1, main_tab2, main_tab3, main_tab_cat, main_tab4 = tabs

# ==========================================
# 1. تبويب البحث الذكي
# ==========================================
with main_tab1:
    search_query = st.text_input(
        "🔍 ابحث هنا بالاسم أو الكود:",
        placeholder="اكتب الكود أو اسم الصنف هنا...",
        key="search_bar"
    )
    
    if search_query:
        q = str(search_query).strip().lower()
        matched = [c for c, v in system_inventory.items() if q in c.lower() or q in v.get('name','').lower()]
        if matched:
            st.markdown("### ✨ نتائج البحث المطابقة:")
            for p_code in matched[:8]:
                item = system_inventory[p_code]
                render_product_card(p_code, item.get('name',''), item.get('sys_stock',0), p_price=item.get('price',0), is_sales=True)
        else:
            st.warning("⚠️ لم يتم العثور على أي منتج يطابق هذا البحث.")

    st.markdown("<br><hr>", unsafe_allow_html=True)
    st.markdown("### 🧠 البحث البصري بالصورة (الذكاء الاصطناعي):")
    cam_tab, up_tab = st.tabs(["📸 التقاط بالكاميرا", "📁 رفع صورة"])
    raw_img = None
    with cam_tab:
        c_photo = st.camera_input("وجّه الكاميرا نحو الحذاء/المنتج", key="vision_cam_input")
        if c_photo: raw_img = Image.open(c_photo).convert("RGB")
    with up_tab:
        u_file = st.file_uploader("ارفع صورة المنتج", type=["jpg", "jpeg", "png"], key="vision_file_uploader")
        if u_file: raw_img = Image.open(u_file).convert("RGB")

    if raw_img:
        st.markdown("### ✂️ قص بؤرة البحث:")
        c_img = st_cropper(raw_img, realtime_update=True, box_color='#1C65A6', aspect_ratio=None)
        if st.button("🚀 ابحث بالذكاء الاصطناعي الآن"):
            with st.spinner('جاري المسح البصري المقارن...'):
                try:
                    emb = get_image_embedding(c_img)
                    res = collection.query(query_embeddings=[emb], n_results=8, include=['distances', 'metadatas'])
                    if res['distances'][0]:
                        u_color = get_color_histogram(c_img)
                        refined = []
                        for i in range(len(res['distances'][0])):
                            fn = res['metadatas'][0][i].get('filename', '')
                            img_p = os.path.join(BASE_DIR, "compressed_images", fn)
                            c_dist = compare_histograms(u_color, get_color_histogram(Image.open(img_p))) if os.path.exists(img_p) else 0
                            refined.append({'fn': fn, 'score': res['distances'][0][i] + (c_dist * 0.5)})
                        refined.sort(key=lambda x: x['score'])
                        for r in refined[:3]:
                            p_code = r['fn'].split('.')[0].upper()
                            if p_code in system_inventory:
                                item = system_inventory[p_code]
                                render_product_card(p_code, item.get('name',''), item.get('sys_stock',0), p_price=item.get('price',0), is_sales=True)
                except Exception as e: st.error(f"⚠️ خطأ أثناء البحث: {e}")

# ==========================================
# 2. تبويب الجرد التشاركي
# ==========================================
with main_tab2:
    shared_inv = load_shared_inventory()
    
    if not shared_inv.get("is_active", False):
        st.markdown("### 🆕 إعداد جلسة جرد جديدة")
        col_i1, col_i2 = st.columns(2)
        with col_i1:
            inv_n = st.text_input("اسم/رقم الجرد (مثال: جرد مخزن رئيسي 1)", key="inv_n_in")
            inv_r = st.selectbox("سبب الجرد", ["جرد دوري", "جرد مفاجئ", "تسليم عهدة", "نهاية العام"], key="inv_r_in")
        with col_i2:
            inv_d = st.date_input("تاريخ الجرد", datetime.date.today(), key="inv_d_in")
            st.write("")
            st.write("")
            if st.button("🚀 فتح جلسة الجرد وتثبيتها", key="open_inv_btn"):
                if not inv_n.strip():
                    st.error("⚠️ اكتب اسم الجرد أولاً للمتابعة!")
                else:
                    save_shared_inventory({
                        "is_active": True,
                        "name": inv_n.strip(),
                        "reason": inv_r,
                        "date": str(inv_d),
                        "scanned_items": {}
                    })
                    st.success("✅ تم فتح الجلسة وتثبيتها بنجاح!")
                    time.sleep(0.3)
                    st.rerun()

        with st.expander("📥 استعادة جلسة جرد سابقة (Backup)", expanded=False):
            backup_file = st.file_uploader("ارفع ملف الجلسة (.json):", type=['json'], key="restore_inv_uploader")
            if backup_file and st.button("استعادة هذه الجلسة فوراً 🔄", key="restore_inv_btn"):
                try:
                    data = json.load(backup_file)
                    data["is_active"] = True
                    save_shared_inventory(data)
                    st.success("✅ تمت استعادة الجلسة بنجاح!")
                    time.sleep(0.5)
                    st.rerun()
                except Exception as e:
                    st.error(f"ملف غير صالح: {e}")
    else:
        st.markdown(f'<div class="inv-active-bar">📌 <b>جلسة جرد نشطة ومحفوظة:</b> {shared_inv.get("name")} | <b>السبب:</b> {shared_inv.get("reason")} | <b>التاريخ:</b> {shared_inv.get("date")}</div>', unsafe_allow_html=True)
        
        tab_sum, tab_scan, tab_edit, tab_rep = st.tabs(["📊 ملخص الأرصدة", "🔫 مسح الباركود", "📝 مراجعة وتعديل المجرد", "⚖️ تقرير الفروقات"])
        scanned_map = shared_inv.get("scanned_items", {})

        with tab_sum:
            total_qty = sum(info.get('sys_stock', 0) for info in system_inventory.values())
            c1, c2, c3 = st.columns(3)
            with c1: st.markdown(f'<div class="metric-card"><div class="metric-title">الأصناف الدفترية</div><div class="metric-value" style="color:#1C65A6;">{len(system_inventory)}</div></div>', unsafe_allow_html=True)
            with c2: st.markdown(f'<div class="metric-card"><div class="metric-title">إجمالي القطع</div><div class="metric-value" style="color:#10B981;">{int(total_qty)}</div></div>', unsafe_allow_html=True)
            with c3: st.markdown(f'<div class="metric-card"><div class="metric-title">إجمالي المجرد فعلياً</div><div class="metric-value" style="color:#F59E0B;">{int(sum(scanned_map.values()))}</div></div>', unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            col_rf, col_bk = st.columns(2)
            with col_rf:
                if st.button("🔄 تحديث البيانات اللحظية", key="ref_inv"): st.rerun()
            with col_bk:
                json_bytes = json.dumps(shared_inv, ensure_ascii=False, indent=4).encode('utf-8')
                st.download_button(
                    "💾 تحميل نسخة احتياطية للجلسة الحالية (JSON)",
                    data=json_bytes,
                    file_name=f"Active_Inventory_{shared_inv.get('name')}.json",
                    mime="application/json"
                )

        with tab_scan:
            if "inv_scan_counter" not in st.session_state:
                st.session_state.inv_scan_counter = 0
            if "current_scanned_code" not in st.session_state:
                st.session_state.current_scanned_code = None

            # الحالة 1: تم مسح كود بنجاح -> إظهار كارت الصنف وخانة الكمية فقط
            if st.session_state.current_scanned_code:
                active_c = st.session_state.current_scanned_code
                item_info = system_inventory[active_c]
                already_counted = scanned_map.get(active_c, 0)
                
                st.markdown(f"<div style='background:#E8F0F8; color:#1C65A6; padding:10px 16px; border-radius:10px; margin-bottom:12px; font-weight:800; font-size:16px;'>🎯 تم لقط الباركود بنجاح! إجمالي المجرد لهذا الموديل: {int(already_counted)} قطعة</div>", unsafe_allow_html=True)
                
                render_product_card(active_c, item_info.get('name', ''), item_info.get('sys_stock', 0))

                with st.form(key=f"inv_form_{active_c}_{st.session_state.inv_scan_counter}", clear_on_submit=True):
                    col_q, col_s = st.columns([3, 2])
                    with col_q:
                        add_q = st.number_input(
                            "🔢 الكمية الفعلية المضافة (اكتب العدد واضغط Enter):",
                            min_value=1,
                            value=1,
                            step=1,
                            key=f"qty_field_{st.session_state.inv_scan_counter}"
                        )
                    with col_s:
                        st.write("")
                        st.write("")
                        confirm_sub = st.form_submit_button("✅ حفظ وإضافة الصنف (Enter)", type="primary", use_container_width=True)

                    if confirm_sub:
                        l_inv = load_shared_inventory()
                        if "scanned_items" not in l_inv:
                            l_inv["scanned_items"] = {}
                        l_inv["scanned_items"][active_c] = max(0, l_inv["scanned_items"].get(active_c, 0) + add_q)
                        save_shared_inventory(l_inv)
                        
                        st.toast(f"✅ تم حفظ إضافة ({add_q}) قطعة للكود [{active_c}]", icon="📦")
                        st.session_state.current_scanned_code = None
                        st.session_state.inv_scan_counter += 1
                        st.rerun()

                focus_qty_js = """
                <script>
                (function() {
                    var tries = 0;
                    var interval = setInterval(function() {
                        tries++;
                        try {
                            var doc = window.parent.document;
                            var qInp = doc.querySelector('input[type="number"]');
                            if (qInp) { qInp.focus(); qInp.select(); clearInterval(interval); }
                        } catch(e) {}
                        if (tries > 30) clearInterval(interval);
                    }, 40);
                })();
                </script>
                """
                components.html(focus_qty_js, height=0, width=0)

                if st.button("⏭️ تخطي والصنف التالي (إلغاء)", key=f"skip_btn_{st.session_state.inv_scan_counter}"):
                    st.session_state.current_scanned_code = None
                    st.session_state.inv_scan_counter += 1
                    st.rerun()

            # الحالة 2: انتظار مسح باركود جديد -> إظهار الإسكانر الحقيقي
            else:
                barcode_field_key = f"barcode_scanner_input_{st.session_state.inv_scan_counter}"
                scanned_raw = st.text_input(
                    "🔫 باركود الصنف (استقبال تلقائي من الكاميرا أو كتابة يدوية):",
                    key=barcode_field_key,
                    placeholder="وجّه الكاميرا نحو الباركود أو مرر الإسكانر..."
                )

                if scanned_raw:
                    matched_k = match_product_code(scanned_raw)
                    if matched_k:
                        st.session_state.current_scanned_code = matched_k
                        st.rerun()
                    else:
                        st.error(f"❌ الباركود '{scanned_raw.strip()}' غير مسجل في قاعدة البيانات.")

                # الكود الأصلي الصاروخي بدون key في components.html
                original_working_scanner_html = """
                <!DOCTYPE html>
                <html lang="ar" dir="rtl">
                <head>
                    <meta charset="utf-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
                    <script src="https://unpkg.com/@zxing/library@latest"></script>
                    <style>
                        body { margin: 0; padding: 4px; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif; text-align: center; background: transparent; }
                        .scanner-box {
                            background: #FFFFFF;
                            border: 2.5px solid #1C65A6;
                            border-radius: 14px;
                            padding: 10px;
                            max-width: 480px;
                            margin: 0 auto;
                            box-shadow: 0 4px 15px rgba(28, 101, 166, 0.15);
                        }
                        .video-container {
                            width: 100%;
                            height: 260px;
                            background: #111111;
                            position: relative;
                        }
                        video {
                            width: 100%;
                            height: 100%;
                            object-fit: cover;
                        }
                        .laser-line {
                            position: absolute;
                            top: 50%;
                            left: 5%;
                            width: 90%;
                            height: 2px;
                            background: #EF4444;
                            box-shadow: 0 0 8px #EF4444;
                            pointer-events: none;
                        }
                        .lens-bar {
                            display: flex;
                            gap: 6px;
                            justify-content: center;
                            flex-wrap: wrap;
                            margin-top: 10px;
                        }
                        .lens-btn {
                            background: #F1F5F9;
                            color: #1E293B;
                            border: 1.5px solid #CBD5E1;
                            border-radius: 8px;
                            padding: 6px 12px;
                            font-size: 13px;
                            font-weight: 700;
                            cursor: pointer;
                        }
                        .lens-btn.active {
                            background: #1C65A6;
                            color: #FFFFFF;
                            border-color: #1C65A6;
                        }
                        #status-bar {
                            margin-top: 8px;
                            font-size: 13.5px;
                            font-weight: 800;
                            color: #1C65A6;
                            min-height: 20px;
                        }
                    </style>
                </head>
                <body>
                    <div class="scanner-box">
                        <div class="video-container">
                            <video id="scanner-feed" autoplay playsinline webkit-playsinline muted></video>
                            <div class="laser-line"></div>
                        </div>
                        
                        <div id="lens-buttons" class="lens-bar">
                            <span style="font-size:12px; color:#64748B;">🔍 جاري تهيئة الكاميرا...</span>
                        </div>

                        <div id="status-bar">⏳ جاري الاتصال بالكاميرا...</div>
                    </div>

                    <script>
                        var activeStream = null;
                        var codeReader = null;
                        var isLocked = false;
                        var detector = null;

                        function playBeep() {
                            try {
                                var ctx = new (window.AudioContext || window.webkitAudioContext)();
                                var osc = ctx.createOscillator();
                                var gain = ctx.createGain();
                                osc.type = "sine";
                                osc.frequency.setValueAtTime(1900, ctx.currentTime);
                                gain.gain.setValueAtTime(0.5, ctx.currentTime);
                                osc.connect(gain);
                                gain.connect(ctx.destination);
                                osc.start();
                                osc.stop(ctx.currentTime + 0.12);
                                if (navigator.vibrate) navigator.vibrate([80, 40, 80]);
                            } catch(e) {}
                        }

                        function sendCode(code) {
                            if (isLocked) return;
                            isLocked = true;
                            playBeep();
                            var cleanCode = code.trim().toUpperCase();
                            document.getElementById("status-bar").innerHTML = "🎯 تم التقاط الصنف: <b>" + cleanCode + "</b>";

                            if (codeReader) { try { codeReader.reset(); } catch(e) {} }
                            if (activeStream) { try { activeStream.getTracks().forEach(t => t.stop()); } catch(e) {} }

                            setTimeout(function() {
                                try {
                                    var url = new URL(window.parent.location.href);
                                    url.searchParams.set("scanned_code", cleanCode);
                                    window.parent.location.href = url.href;
                                } catch(e) {}
                            }, 120);
                        }

                        async function initDetector() {
                            if ('BarcodeDetector' in window) {
                                try {
                                    detector = new BarcodeDetector({ formats: ['code_128', 'code_39', 'ean_13', 'ean_8', 'upc_a', 'upc_e', 'qr_code'] });
                                } catch(e) { detector = null; }
                            }
                        }

                        function scanLoop(videoEl) {
                            if (isLocked) return;
                            if (detector && videoEl.readyState >= 2) {
                                detector.detect(videoEl).then(barcodes => {
                                    if (barcodes.length > 0 && barcodes[0].rawValue) {
                                        sendCode(barcodes[0].rawValue.trim());
                                        return;
                                    }
                                    if (!isLocked) requestAnimationFrame(() => scanLoop(videoEl));
                                }).catch(() => {
                                    if (!isLocked) requestAnimationFrame(() => scanLoop(videoEl));
                                });
                            } else {
                                if (!isLocked) requestAnimationFrame(() => scanLoop(videoEl));
                            }
                        }

                        async function startLens(deviceId, btnElement) {
                            var statusEl = document.getElementById("status-bar");
                            statusEl.innerHTML = "⏳ جاري تشغيل العدسة...";

                            document.querySelectorAll('.lens-btn').forEach(b => b.classList.remove('active'));
                            if (btnElement) btnElement.classList.add('active');

                            if (activeStream) {
                                activeStream.getTracks().forEach(t => t.stop());
                                activeStream = null;
                            }
                            if (codeReader) { try { codeReader.reset(); } catch(e) {} }

                            var videoConstraints = { width: { ideal: 1280 }, height: { ideal: 720 } };
                            if (deviceId) videoConstraints.deviceId = { exact: deviceId };
                            else videoConstraints.facingMode = { ideal: "environment" };

                            try {
                                const stream = await navigator.mediaDevices.getUserMedia({ audio: false, video: videoConstraints });
                                activeStream = stream;
                                var v = document.getElementById("scanner-feed");
                                v.srcObject = stream;
                                v.setAttribute('playsinline', 'true');
                                v.setAttribute('webkit-playsinline', 'true');
                                v.muted = true;
                                await v.play();
                                statusEl.innerHTML = "🟢 الكاميرا تعمل - وجّه الباركود داخل الإطار";

                                if (detector) {
                                    scanLoop(v);
                                } else {
                                    codeReader = new ZXing.BrowserMultiFormatReader();
                                    codeReader.decodeFromVideoElement(v, (result, err) => {
                                        if (result && result.text) sendCode(result.text);
                                    });
                                }
                            } catch(err) {
                                statusEl.innerHTML = "⚠️ تعذر تشغيل هذه العدسة (جرب الضغط على عدسة أخرى أعلاه).";
                            }
                        }

                        async function setupCameras() {
                            await initDetector();
                            try {
                                const promptStream = await navigator.mediaDevices.getUserMedia({ video: true, audio: false });
                                promptStream.getTracks().forEach(t => t.stop());

                                const devices = await navigator.mediaDevices.enumerateDevices();
                                const videoDevices = devices.filter(d => d.kind === 'videoinput');
                                const lensContainer = document.getElementById("lens-buttons");
                                lensContainer.innerHTML = "";

                                if (videoDevices.length === 0) {
                                    document.getElementById("status-bar").innerHTML = "❌ لم يتم العثور على أي كاميرا!";
                                    return;
                                }

                                let targetIndex = 0;
                                videoDevices.forEach((dev, idx) => {
                                    const b = document.createElement("button");
                                    b.className = "lens-btn";
                                    const lbl = (dev.label || ("عدسة " + (idx + 1))).toLowerCase();
                                    const isBack = lbl.includes('back') || lbl.includes('rear') || lbl.includes('environment') || lbl.includes('خلف');
                                    
                                    b.innerText = (isBack ? "📸 خلفية " : "🤳 أمامية ") + (idx + 1);
                                    b.onclick = function() { startLens(dev.deviceId, b); };
                                    lensContainer.appendChild(b);

                                    if (isBack) targetIndex = idx;
                                });

                                var firstBtn = lensContainer.children[targetIndex] || lensContainer.children[0];
                                startLens(videoDevices[targetIndex].deviceId, firstBtn);
                            } catch(e) {
                                startLens(null, null);
                            }
                        }

                        window.addEventListener('load', () => setTimeout(setupCameras, 200));
                    </script>
                </body>
                </html>
                """
                components.html(original_working_scanner_html, height=440)

        # 📝 مراجعة وتعديل الجلسة
        with tab_edit:
            st.markdown("### 📝 قائمة الأصناف المجرودة بالجلسة الحالية (تعديل وحذف مباشر)")
            if scanned_map:
                active_list = []
                for c, q in scanned_map.items():
                    active_list.append({
                        "كود الصنف": c,
                        "اسم الصنف": system_inventory.get(c, {}).get('name', 'غير مسجل'),
                        "الكمية المجردة": int(q)
                    })
                
                df_active = pd.DataFrame(active_list)
                edited_df = st.data_editor(
                    df_active,
                    column_config={
                        "كود الصنف": st.column_config.TextColumn("كود الصنف", disabled=True),
                        "اسم الصنف": st.column_config.TextColumn("اسم الصنف", disabled=True),
                        "الكمية المجردة": st.column_config.NumberColumn("الكمية الفعلية المجردة", min_value=0, step=1, required=True)
                    },
                    num_rows="dynamic",
                    use_container_width=True,
                    key="active_session_data_editor"
                )
                
                if st.button("💾 تطبيق وحفظ تعديلات الجلسة", type="primary", key="save_edited_session_btn"):
                    new_scanned_map = {}
                    for _, r in edited_df.iterrows():
                        c_k = str(r["كود الصنف"]).strip().upper()
                        q_v = int(r["الكمية المجردة"])
                        if c_k and q_v > 0:
                            new_scanned_map[c_k] = q_v
                    
                    l_inv = load_shared_inventory()
                    l_inv["scanned_items"] = new_scanned_map
                    save_shared_inventory(l_inv)
                    st.success("🎉 تم تحديث وحفظ بيانات الجلسة بنجاح!")
                    time.sleep(0.4)
                    st.rerun()
            else:
                st.warning("⚠️ لم يتم مسح أي صنف في هذه الجلسة حتى الآن.")

        # ⚖️ تقرير الفروقات
        with tab_rep:
            rep_data = [
                {
                    "كود الصنف": c,
                    "اسم الصنف": i.get('name', ''),
                    "الرصيد الدفتري": i.get('sys_stock', 0),
                    "الرصيد الفعلي": scanned_map.get(c, 0),
                    "الفروقات": scanned_map.get(c, 0) - i.get('sys_stock', 0)
                }
                for c, i in system_inventory.items()
            ]
            
            filter_rep = st.text_input("🔍 تصفية سريعة بتقرير الفروقات:", placeholder="ابحث بكود أو اسم الموديل...", key="rep_filter_input")
            filtered_rep = rep_data
            if filter_rep:
                fr = filter_rep.strip().lower()
                filtered_rep = [r for r in rep_data if fr in str(r.get("كود الصنف", "")).lower() or fr in str(r.get("اسم الصنف", "")).lower()]
            
            df_rep = pd.DataFrame(filtered_rep)
            st.dataframe(df_rep, use_container_width=True, hide_index=True)
            
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w: df_rep.to_excel(w, index=False)
            st.download_button("📥 تحميل تقرير الجرد (Excel)", buf.getvalue(), f"Inventory_{shared_inv.get('name')}.xlsx")
            
            st.markdown("---")
            if st.session_state.current_user == "abobakr":
                if st.button("🛑 إغلاق وإنهاء جلسة الجرد نهائياً (ترحيل للأرشيف)", type="primary", key="close_inv_session"):
                    save_to_inv_history({
                        "timestamp": str(datetime.datetime.now()),
                        "name": shared_inv.get('name'),
                        "date": shared_inv.get('date'),
                        "report": rep_data
                    })
                    save_shared_inventory({"is_active": False, "scanned_items": {}})
                    st.success("✅ تم إنهاء الجلسة وحفظها في الأرشيف بنجاح!")
                    time.sleep(0.5)
                    st.rerun()

# ==========================================
# 3. تبويب فواتير الجملة
# ==========================================
with main_tab3:
    shared_sales = load_shared_sales()
    
    if not shared_sales.get("is_active", False):
        st.markdown("### 🏬 فتح وردية مبيعات جملة جديدة")
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            s_name = st.text_input("اسم/رقم الوردية (مثال: وردية صباحية 1)", key="s_name_in")
        with col_s2:
            s_date = st.date_input("تاريخ الوردية", datetime.date.today(), key="s_date_in")
            st.write("")
            st.write("")
            if st.button("🚀 فتح وردية البيع وتثبيتها", key="open_sales_btn"):
                if not s_name.strip(): st.error("⚠️ اكتب اسم الوردية أولاً!")
                else:
                    save_shared_sales({"is_active": True, "name": s_name.strip(), "date": str(s_date), "invoices": [], "deductions": {}})
                    st.rerun()
    else:
        st.markdown(f'<div class="sales-active-bar">💳 <b>وردية الجملة النشطة:</b> {shared_sales.get("name")} | <b>التاريخ:</b> {shared_sales.get("date")}</div>', unsafe_allow_html=True)
        
        all_invs = shared_sales.get("invoices", [])
        total_rev = sum(float(inv.get("total", 0.0)) for inv in all_invs)
        my_rev = sum(float(inv.get("total", 0.0)) for inv in all_invs if inv.get("salesperson") == st.session_state.current_user)
        
        c1, c2, c3 = st.columns(3)
        with c1: st.markdown(f'<div class="metric-card"><div class="metric-title">إجمالي إيراد الوردية</div><div class="metric-value" style="color:#1C65A6;">{total_rev:.2f} ج.م</div></div>', unsafe_allow_html=True)
        with c2: st.markdown(f'<div class="metric-card"><div class="metric-title">مبيعاتي بالوردية</div><div class="metric-value" style="color:#F59E0B;">{my_rev:.2f} ج.م</div></div>', unsafe_allow_html=True)
        with c3: st.markdown(f'<div class="metric-card"><div class="metric-title">عدد الفواتير</div><div class="metric-value" style="color:#10B981;">{len(all_invs)}</div></div>', unsafe_allow_html=True)
        st.markdown("---")

        if "active_cust" not in st.session_state: st.session_state.active_cust = ""
        if "cart" not in st.session_state: st.session_state.cart = []
        if "pos_scan_counter" not in st.session_state: st.session_state.pos_scan_counter = 0

        if not st.session_state.active_cust:
            st.markdown("### 📝 فتح فاتورة لعميل جديد")
            c_name_input = st.text_input("اسم العميل / المحل:", key="cust_name_field")
            if st.button("بدء الفاتورة 🛒", key="start_inv_btn"):
                if c_name_input.strip():
                    st.session_state.active_cust = c_name_input.strip()
                    st.session_state.cart = []
                    st.rerun()
                else: st.error("⚠️ اكتب اسم العميل للمتابعة.")
        else:
            st.markdown(f"### 🧾 العميل الحالي: <span style='color:#1C65A6;'>{st.session_state.active_cust}</span>", unsafe_allow_html=True)
            
            pos_code_key = f"pos_barcode_in_{st.session_state.pos_scan_counter}"
            scan_pos = st.text_input(
                "🔫 باركود الصنف للفاتورة:",
                key=pos_code_key,
                placeholder="امسح باركود الصنف للفاتورة..."
            )
            
            matched_pos_code = match_product_code(scan_pos) if scan_pos else None

            if matched_pos_code:
                p_c = matched_pos_code
                p_data = system_inventory[p_c]
                s_qty = float(p_data.get('sys_stock', 0.0))
                price = float(p_data.get('price', 0.0))
                sold_q = float(shared_sales.get("deductions", {}).get(p_c, 0.0))
                in_cart = sum(float(it.get('qty',0)) for it in st.session_state.cart if it.get('code') == p_c)
                avail = s_qty - sold_q - in_cart
                
                render_product_card(p_c, p_data.get('name',''), avail, p_price=price, is_sales=True)
                
                if avail <= 0:
                    st.error("❌ تحذير: هذا الصنف غير متاح في المخزن حالياً (رصيده 0)!")
                else:
                    with st.form(key=f"pos_add_form_{p_c}_{st.session_state.pos_scan_counter}", clear_on_submit=True):
                        c_q1, c_q2, c_q3 = st.columns([3, 2, 2])
                        with c_q1:
                            req_qty = st.number_input("الكمية المطلوبة (Enter للحفظ السريع):", min_value=1, max_value=int(avail), value=1, key=f"pos_qty_{st.session_state.pos_scan_counter}")
                        with c_q2:
                            st.number_input("السعر المسجل:", value=price, disabled=True, key=f"pos_p_{st.session_state.pos_scan_counter}")
                        with c_q3:
                            st.write("")
                            st.write("")
                            add_cart_sub = st.form_submit_button("📥 إضافة للفاتورة (Enter)", type="primary", use_container_width=True)

                        if add_cart_sub:
                            if price <= 0:
                                st.error("⚠️ الصنف ليس له سعر مسجل في النظام.")
                            else:
                                st.session_state.cart.append({
                                    "code": p_c,
                                    "name": p_data.get('name',''),
                                    "qty": req_qty,
                                    "price": price,
                                    "total": req_qty * price
                                })
                                st.toast(f"✅ تمت إضافة ({req_qty}) قطعة من [{p_c}] للفاتورة", icon="🛒")
                                st.session_state.pos_scan_counter += 1
                                st.rerun()

                    focus_pos_qty_js = """
                    <script>
                    (function() {
                        var tries = 0;
                        var interval = setInterval(function() {
                            tries++;
                            try {
                                var doc = window.parent.document;
                                var qInp = doc.querySelector('input[type="number"]');
                                if (qInp) { qInp.focus(); qInp.select(); clearInterval(interval); }
                            } catch(e) {}
                            if (tries > 30) clearInterval(interval);
                        }, 40);
                    })();
                    </script>
                    """
                    components.html(focus_pos_qty_js, height=0, width=0)

            elif scan_pos:
                st.error(f"❌ الكود '{scan_pos.strip()}' غير مسجل في النظام.")

            if st.session_state.cart:
                st.markdown("<br>#### 🛒 الأصناف المضافة بالفاتورة حتى الآن:", unsafe_allow_html=True)
                df_c = pd.DataFrame(st.session_state.cart)
                st.dataframe(df_c[['code', 'name', 'qty', 'price', 'total']], use_container_width=True)
                cart_tot = sum(float(it.get('total', 0.0)) for it in st.session_state.cart)
                st.markdown(f"<h3 style='color:#DC2626;'>الإجمالي المطلوب: {cart_tot:.2f} ج.م</h3>", unsafe_allow_html=True)
                
                col_sv, col_cn = st.columns(2)
                with col_sv:
                    if st.button("✅ اعتماد الفاتورة وتحميل الإكسيل", type="primary", key="save_inv_btn"):
                        l_sales = load_shared_sales()
                        new_id = max([i.get("invoice_id", 0) for i in l_sales.get("invoices", [])] + [0]) + 1
                        l_sales["invoices"].append({
                            "invoice_id": new_id, "time": str(datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
                            "salesperson": st.session_state.current_user, "customer": st.session_state.active_cust,
                            "total": cart_tot, "items": st.session_state.cart
                        })
                        for it in st.session_state.cart:
                            l_sales["deductions"][it['code']] = l_sales.get("deductions", {}).get(it['code'], 0) + it['qty']
                        save_shared_sales(l_sales)
                        
                        df_ex = pd.DataFrame(st.session_state.cart).rename(columns={'code':'كود الصنف', 'name':'اسم الصنف', 'qty':'الكمية', 'price':'السعر', 'total':'الإجمالي'})
                        buf = io.BytesIO()
                        with pd.ExcelWriter(buf, engine='openpyxl') as w: df_ex.to_excel(w, index=False)
                        st.download_button("📥 تحميل الفاتورة (Excel)", buf.getvalue(), f"Invoice_{new_id}_{st.session_state.active_cust}.xlsx")
                        
                        st.session_state.active_cust = ""
                        st.session_state.cart = []
                        st.session_state.pos_scan_counter += 1
                        st.success("🎉 تم اعتماد الفاتورة وخصم الكميات بنجاح!")
                with col_cn:
                    if st.button("🗑️ إلغاء الفاتورة بالكامل", key="cancel_inv_btn"):
                        st.session_state.active_cust = ""
                        st.session_state.cart = []
                        st.session_state.pos_scan_counter += 1
                        st.rerun()

        st.markdown("---")
        if st.session_state.current_user == "abobakr":
            if st.button("🛑 إغلاق وردية البيع نهائياً (ترحيل للأرشيف)", key="close_sales_shift_btn"):
                save_to_sales_history({"name": shared_sales.get("name"), "date": shared_sales.get("date"), "total_revenue": total_rev, "invoices": all_invs})
                save_shared_sales({"is_active": False, "invoices": [], "deductions": {}})
                st.success("✅ تم ترحيل الوردية للأرشيف وإغلاقها.")
                time.sleep(0.5)
                st.rerun()

# ==========================================
# 4. تبويب الكاتالوج الشامل
# ==========================================
with main_tab_cat:
    st.markdown("### 📖 الكاتالوج الشامل للأصناف المتوفرة (Live Catalog)")
    shared_s_cat = load_shared_sales()
    deductions = shared_s_cat.get("deductions", {})
    
    cat_rows = []
    for p_c, p_inf in system_inventory.items():
        stk_before = float(p_inf.get('sys_stock', 0.0))
        sold_amt = float(deductions.get(p_c, 0.0))
        stk_avail = stk_before - sold_amt
        price_val = float(p_inf.get('price', 0.0))
        
        if stk_avail > 0:
            thumb_val = None
            raw_img_path = None
            for ext in ['.jpg', '.jpeg', '.png', '.JPG']:
                img_p = os.path.join(BASE_DIR, "compressed_images", f"{p_c}{ext}")
                if os.path.exists(img_p):
                    raw_img_path = img_p
                    thumb_val = get_thumbnail_base64(img_p)
                    break
                    
            cat_rows.append({
                "صورة المنتج": thumb_val,
                "كود الصنف": p_c,
                "اسم الصنف": p_inf.get('name', ''),
                "سعر القطعة (ج.م)": price_val,
                "الرصيد الدفتري (قبل البيع)": stk_before,
                "كمية المبيعات بالوردية": sold_amt,
                "الرصيد اللحظي المتاح": stk_avail,
                "img_path": raw_img_path
            })
            
    if cat_rows:
        cat_rows.sort(key=lambda x: str(x.get("كود الصنف", "")).upper())
        filter_q = st.text_input("🔍 تصفية سريعة بالكاتالوج:", placeholder="ابحث بكود أو اسم الموديل...", key="cat_filter")
        
        filtered_rows = cat_rows
        if filter_q:
            fq = filter_q.strip().lower()
            filtered_rows = [r for r in cat_rows if fq in str(r.get("كود الصنف","")).lower() or fq in str(r.get("اسم الصنف","")).lower()]
            
        df_display = pd.DataFrame(filtered_rows).drop(columns=["img_path"], errors="ignore")
        
        st.dataframe(
            df_display,
            column_config={
                "صورة المنتج": st.column_config.ImageColumn("صورة المنتج", help="صورة الصنف واضحة"),
                "كود الصنف": st.column_config.TextColumn("كود الصنف"),
                "اسم الصنف": st.column_config.TextColumn("اسم الصنف"),
                "سعر القطعة (ج.م)": st.column_config.NumberColumn("سعر القطعة (ج.م)", format="%.2f"),
                "الرصيد الدفتري (قبل البيع)": st.column_config.NumberColumn("الرصيد الدفتري"),
                "كمية المبيعات بالوردية": st.column_config.NumberColumn("المبيعات"),
                "الرصيد اللحظي المتاح": st.column_config.NumberColumn("الرصيد المتاح للبيع")
            },
            use_container_width=True,
            hide_index=True
        )
        
        excel_catalog_bytes = generate_catalog_excel(filtered_rows)
        st.download_button(
            label="📥 تحميل الكاتالوج الكامل بالصور (Excel)",
            data=excel_catalog_bytes,
            file_name=f"Live_Catalog_Images_{datetime.date.today()}.xlsx",
            type="primary"
        )
    else:
        st.info("📦 لا توجد أي أصناف متوفرة في المستودع حالياً (الأرصدة صفر).")

# ==========================================
# 5. تبويب لوحة تحكم الإدارة
# ==========================================
with main_tab4:
    st.markdown("## 📈 لوحة تحكم الإدارة (Live Dashboard)")
    master_sales = []
    for rec in load_sales_history():
        for inv in rec.get('invoices', []):
            for it in inv.get('items', []):
                master_sales.append({"الوردية": rec.get('name', ''), "العميل": inv.get('customer', ''), "البائع": inv.get('salesperson', ''), "كود الصنف": it.get('code', ''), "الكمية": float(it.get('qty', 0)), "الإجمالي": float(it.get('total', 0.0))})
        for inv in load_shared_sales().get('invoices', []):
            for it in inv.get('items', []):
                master_sales.append({"الوردية": "نشطة حالياً", "العميل": inv.get('customer', ''), "البائع": inv.get('salesperson', ''), "كود الصنف": it.get('code', ''), "الكمية": float(it.get('qty', 0)), "الإجمالي": float(it.get('total', 0.0))})
                
        if master_sales:
            df_all = pd.DataFrame(master_sales)
            tot_cash = df_all['الإجمالي'].sum()
            tot_pieces = df_all['الكمية'].sum()
            
            c_a1, c_a2 = st.columns(2)
            with c_a1: st.markdown(f'<div class="metric-card"><div class="metric-title">💰 إجمالي مبيعات المحل</div><div class="metric-value" style="color:#10B981;">{tot_cash:.2f} ج.م</div></div>', unsafe_allow_html=True)
            with c_a2: st.markdown(f'<div class="metric-card"><div class="metric-title">📦 إجمالي القطع المباعة</div><div class="metric-value" style="color:#1C65A6;">{int(tot_pieces)} قطعة</div></div>', unsafe_allow_html=True)
            st.markdown("---")
            
            st.markdown("#### 👥 تقرير مشتريات العملاء:")
            cust_summary = df_all.groupby('العميل').agg({'الكمية':'sum', 'الإجمالي':'sum'}).reset_index().sort_values(by='الإجمالي', ascending=False)
            st.dataframe(cust_summary, use_container_width=True, hide_index=True)
            
            buf_all = io.BytesIO()
            with pd.ExcelWriter(buf_all, engine='openpyxl') as w: df_all.to_excel(w, index=False)
            st.download_button("📥 تحميل كل سجلات المبيعات (Excel)", buf_all.getvalue(), f"All_Sales_{datetime.date.today()}.xlsx")
        else: st.info("لا توجد مبيعات مسجلة حتى الآن.")

st.markdown('<div class="footer">تصميم وبرمجة: <span>أبوبكر عادل</span> © 2026</div>', unsafe_allow_html=True)